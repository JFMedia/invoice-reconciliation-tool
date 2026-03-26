import os
import pandas as pd
from rapidfuzz import fuzz
from extract_invoice import extract_invoice_data
from load_po import load_po_csv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def highlight_excel_report(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    red_fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
    green_fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")

    headers = [cell.value for cell in ws[1]]
    col_map = {header: idx + 1 for idx, header in enumerate(headers)}

    status_col = col_map["status"]
    po_qty_col = col_map["po_qty"]
    invoice_qty_col = col_map["invoice_qty"]
    po_cost_col = col_map["po_unit_cost"]
    invoice_cost_col = col_map["invoice_unit_cost"]

    for row in range(2, ws.max_row + 1):
        status_value = ws.cell(row=row, column=status_col).value

        if status_value == "QTY_MISMATCH":
            ws.cell(row=row, column=po_qty_col).fill = red_fill
            ws.cell(row=row, column=invoice_qty_col).fill = green_fill

        elif status_value == "COST_MISMATCH":
            ws.cell(row=row, column=po_cost_col).fill = red_fill
            ws.cell(row=row, column=invoice_cost_col).fill = green_fill

        elif status_value == "QTY_AND_COST_MISMATCH":
            ws.cell(row=row, column=po_qty_col).fill = red_fill
            ws.cell(row=row, column=invoice_qty_col).fill = green_fill
            ws.cell(row=row, column=po_cost_col).fill = red_fill
            ws.cell(row=row, column=invoice_cost_col).fill = green_fill

        elif status_value == "LINE_MISSING_FROM_PO":
            ws.cell(row=row, column=invoice_qty_col).fill = red_fill
            ws.cell(row=row, column=invoice_cost_col).fill = red_fill

        elif status_value == "LINE_MISSING_FROM_INVOICE":
            ws.cell(row=row, column=po_qty_col).fill = red_fill
            ws.cell(row=row, column=po_cost_col).fill = red_fill

    wb.save(file_path)


def clean_sku(value):
    return "".join(ch for ch in str(value).strip().upper() if ch.isalnum())


def clean_desc(value):
    return " ".join(str(value).strip().upper().split())


def normalize_invoice_items(invoice_data):
    rows = []

    for item in invoice_data.get("items", []):
        rows.append({
            "vendor": str(invoice_data.get("vendor", "")).strip(),
            "invoice_number": str(invoice_data.get("invoice_number", "")).strip(),
            "po_number": str(invoice_data.get("po_number", "")).strip(),
            "po_number_clean": str(invoice_data.get("po_number_clean", "")).strip(),
            "vendor_id": str(item.get("vendor_id", "")).strip(),
            "sku": clean_sku(item.get("sku", "")),
            "description": str(item.get("description", "")).strip(),
            "qty": float(item.get("quantity", 0) or 0),
            "unit_cost": float(item.get("unit_cost", 0) or 0),
            "line_total": float(item.get("line_total", 0) or 0),
        })

    return pd.DataFrame(rows)


def load_all_invoices_for_batch(batch_folder):
    all_rows = []

    for filename in os.listdir(batch_folder):
        if not filename.lower().endswith(".pdf"):
            continue

        pdf_path = os.path.join(batch_folder, filename)
        print(f"Extracting invoice: {pdf_path}")

        invoice_data = extract_invoice_data(pdf_path)
        invoice_df = normalize_invoice_items(invoice_data)

        if not invoice_df.empty:
            all_rows.append(invoice_df)

    if not all_rows:
        return pd.DataFrame()

    return pd.concat(all_rows, ignore_index=True)


def combine_invoice_lines(invoice_df):
    if invoice_df.empty:
        return invoice_df

    invoice_df = invoice_df.copy()

    invoice_df["match_key"] = invoice_df.apply(
        lambda row: row["sku"] if row["sku"] else clean_desc(row["description"]),
        axis=1
    )

    grouped = (
        invoice_df.groupby(
            ["vendor", "po_number", "po_number_clean", "match_key", "sku", "description"],
            dropna=False,
            as_index=False
        )
        .agg({
            "vendor_id": "first",
            "invoice_number": lambda x: ", ".join(
                sorted(set(str(v).strip() for v in x if str(v).strip()))
            ),
            "qty": "sum",
            "line_total": "sum",
            "unit_cost": "mean"
        })
    )

    return grouped


def find_best_match(inv_row, po_df):
    inv_sku = clean_sku(inv_row.get("sku", ""))
    inv_vendor_id = clean_sku(inv_row.get("vendor_id", ""))
    inv_desc = clean_desc(inv_row.get("description", ""))

    # 1. Exact cleaned SKU match
    if inv_sku:
        exact = po_df[po_df["sku"].apply(clean_sku) == inv_sku]
        if not exact.empty:
            return exact.iloc[0]

    # 2. Exact vendor ID match
    if inv_vendor_id:
        vendor_matches = po_df[po_df["vendor_id"].apply(clean_sku) == inv_vendor_id]
        if not vendor_matches.empty:
            return vendor_matches.iloc[0]

    # 3. Exact cleaned description match
    if inv_desc:
        desc_exact = po_df[
            po_df["description"].apply(clean_desc) == inv_desc
        ]
        if not desc_exact.empty:
            return desc_exact.iloc[0]

    # 4. Fuzzy description match
    best_score = 0
    best_row = None

    for _, po_row in po_df.iterrows():
        po_desc = clean_desc(po_row.get("description", ""))
        score = fuzz.token_set_ratio(inv_desc, po_desc)

        if score > best_score:
            best_score = score
            best_row = po_row

    if best_score >= 85:
        return best_row

    return None


def compare_invoice_to_po(invoice_df, po_df):
    results = []
    matched_po_indexes = set()

    vendor_value = ""
    po_number_value = ""

    if not invoice_df.empty:
        vendor_value = str(invoice_df.iloc[0].get("vendor", "")).strip()
        po_number_value = str(invoice_df.iloc[0].get("po_number", "")).strip()

    for _, inv_row in invoice_df.iterrows():
        available_po = po_df.drop(index=matched_po_indexes, errors="ignore")
        match = find_best_match(inv_row, available_po)

        if match is None:
            results.append({
                "vendor": inv_row.get("vendor", "") or vendor_value,
                "po_number": inv_row.get("po_number", "") or po_number_value,
                "invoice_number": inv_row.get("invoice_number", ""),
                "vendor_id": inv_row.get("vendor_id", ""),
                "sku": inv_row.get("sku", ""),
                "description": inv_row.get("description", ""),
                "status": "LINE_MISSING_FROM_PO",
                "po_qty": "",
                "invoice_qty": inv_row.get("qty", ""),
                "qty_difference": "",
                "po_unit_cost": "",
                "invoice_unit_cost": inv_row.get("unit_cost", ""),
                "cost_difference": "",
            })
            continue

        matched_po_indexes.add(match.name)

        qty_diff = round(float(inv_row["qty"]) - float(match["qty"]), 2)
        cost_diff = round(float(inv_row["unit_cost"]) - float(match["unit_cost"]), 2)

        qty_match = qty_diff == 0
        cost_match = cost_diff == 0

        if qty_match and cost_match:
            status = "MATCH"
        elif not qty_match and cost_match:
            status = "QTY_MISMATCH"
        elif qty_match and not cost_match:
            status = "COST_MISMATCH"
        else:
            status = "QTY_AND_COST_MISMATCH"

        results.append({
            "vendor": inv_row.get("vendor", "") or vendor_value,
            "po_number": inv_row.get("po_number", "") or match.get("po_number", po_number_value),
            "invoice_number": inv_row.get("invoice_number", ""),
            "vendor_id": inv_row.get("vendor_id", "") or match.get("vendor_id", ""),
            "sku": inv_row.get("sku", "") or match.get("sku", ""),
            "description": inv_row.get("description", "") or match.get("description", ""),
            "status": status,
            "po_qty": match.get("qty", ""),
            "invoice_qty": inv_row.get("qty", ""),
            "qty_difference": qty_diff,
            "po_unit_cost": match.get("unit_cost", ""),
            "invoice_unit_cost": inv_row.get("unit_cost", ""),
            "cost_difference": cost_diff,
        })

    for idx, po_row in po_df.iterrows():
        if idx not in matched_po_indexes:
            results.append({
                "vendor": vendor_value,
                "po_number": po_row.get("po_number", "") or po_number_value,
                "invoice_number": "",
                "vendor_id": po_row.get("vendor_id", ""),
                "sku": po_row.get("sku", ""),
                "description": po_row.get("description", ""),
                "status": "LINE_MISSING_FROM_INVOICE",
                "po_qty": po_row.get("qty", ""),
                "invoice_qty": "",
                "qty_difference": "",
                "po_unit_cost": po_row.get("unit_cost", ""),
                "invoice_unit_cost": "",
                "cost_difference": "",
            })

    return pd.DataFrame(results)